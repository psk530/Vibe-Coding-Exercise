"""API + static file server for the US TV Market brand dashboard.

Serves brand-dashboard.html and two read-only JSON endpoints backed by
db/dashboard.db (built by etl/load_excel_to_db.py):

  GET /api/weekly  -> every (week, brand) row: the raw fact table
  GET /api/totals  -> all-time per-brand rollup (rev/units/traffic/conv)

The frontend fetches these once on load and does its own period/brand
filtering client-side, exactly as it did with the old embedded JS arrays --
only the data source changed.

Run with:
    uvicorn server.main:app --reload --port 8000
then open http://localhost:8000/
"""
import html
import io
import json
import os
import re
import sqlite3
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime
from pathlib import Path
from urllib.parse import quote

import bcrypt
import httpx
import psycopg2
import psycopg2.extras
from fastapi import Depends, FastAPI, HTTPException, Request
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from openai import OpenAI
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pydantic import BaseModel
from starlette.middleware.sessions import SessionMiddleware

ROOT = Path(__file__).resolve().parent.parent
DB_PATH = ROOT / "db" / "dashboard.db"

app = FastAPI(title="Brand Dashboard API")
app.add_middleware(
    SessionMiddleware,
    secret_key=os.environ.get("SESSION_SECRET_KEY", "dev-insecure-secret-change-me"),
    same_site="lax",
)

CHAT_MODEL = os.environ.get("OPENROUTER_MODEL", "openai/gpt-4o")

NAVER_SEARCH_URL = "https://naverapihub.apigw.ntruss.com/search/v1/{kind}"
NAVER_SEARCH_TRIGGERS = ["최신", "동향", "트렌드", "요즘", "최근", "뉴스", "신제품", "출시", "프로모션", "이벤트", "왜", "원인", "이유"]


def needs_external_search(message: str) -> bool:
    return any(kw in message for kw in NAVER_SEARCH_TRIGGERS)


def _clean_naver_text(s: str) -> str:
    return html.unescape(re.sub(r"</?b>", "", s or ""))


def naver_search(query: str, kind: str, display: int = 3) -> list[dict]:
    client_id = os.environ.get("NAVER_CLIENT_ID")
    client_secret = os.environ.get("NAVER_CLIENT_SECRET")
    if not client_id or not client_secret:
        return []
    try:
        resp = httpx.get(
            NAVER_SEARCH_URL.format(kind=kind),
            params={"query": query, "display": display, "sort": "date"},
            headers={
                "X-NCP-APIGW-API-KEY-ID": client_id,
                "X-NCP-APIGW-API-KEY": client_secret,
            },
            timeout=3,
        )
        resp.raise_for_status()
        return resp.json().get("items", [])
    except Exception:
        return []


def build_naver_context(message: str) -> str:
    # News and blog are fetched concurrently -- sequential calls could add up to
    # 2x the per-request timeout on top of the LLM call, and this endpoint was
    # observed taking ~40s end-to-end (right at the edge of Render/Cloudflare's
    # proxy timeout, occasionally tipping over into a 502).
    with ThreadPoolExecutor(max_workers=2) as pool:
        news_future = pool.submit(naver_search, message, "news", 3)
        blog_future = pool.submit(naver_search, message, "blog", 2)
        items = news_future.result() + blog_future.result()
    if not items:
        return ""
    lines = []
    for item in items:
        title = _clean_naver_text(item.get("title", ""))
        desc = _clean_naver_text(item.get("description", ""))
        link = item.get("link", "")
        lines.append(f"- {title}: {desc} ({link})")
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Auth / users (Postgres -- NOT the local SQLite file: Render's disk is wiped
# on every deploy, so account data needs to live in an external DB to survive
# the auto-deploy-on-every-push workflow this project uses).
# ---------------------------------------------------------------------------

def get_users_conn():
    dsn = os.environ.get("DATABASE_URL")
    if not dsn:
        raise HTTPException(500, "DATABASE_URL이 서버에 설정되어 있지 않습니다.")
    return psycopg2.connect(dsn)


@app.on_event("startup")
def init_users_table():
    # Best-effort only: a Postgres outage (e.g. a paused free-tier Supabase
    # project) must NOT take down the rest of the dashboard, which doesn't
    # depend on it. Auth endpoints will just 500 individually until it's back.
    dsn = os.environ.get("DATABASE_URL")
    if not dsn:
        print("DATABASE_URL not set -- skipping users table init (auth endpoints will 500 until it's configured)")
        return
    try:
        conn = psycopg2.connect(dsn, connect_timeout=5)
        try:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    CREATE TABLE IF NOT EXISTS users (
                        id SERIAL PRIMARY KEY,
                        name TEXT NOT NULL,
                        email TEXT UNIQUE NOT NULL,
                        password_hash TEXT NOT NULL,
                        role TEXT NOT NULL DEFAULT 'user',
                        status TEXT NOT NULL DEFAULT 'active',
                        created_at TIMESTAMPTZ NOT NULL DEFAULT now(),
                        status_changed_at TIMESTAMPTZ NOT NULL DEFAULT now()
                    )
                    """
                )
            conn.commit()
        finally:
            conn.close()
    except Exception as e:
        print(f"WARNING: could not initialize users table (auth will be unavailable until this is fixed): {e!r}")


def hash_password(pw: str) -> str:
    return bcrypt.hashpw(pw.encode(), bcrypt.gensalt()).decode()


def verify_password(pw: str, hashed: str) -> bool:
    try:
        return bcrypt.checkpw(pw.encode(), hashed.encode())
    except Exception:
        return False


PUBLIC_USER_FIELDS = "id, name, email, role, status, created_at, status_changed_at"


def get_current_user(request: Request) -> dict:
    user_id = request.session.get("user_id")
    if not user_id:
        raise HTTPException(401, "로그인이 필요합니다.")
    conn = get_users_conn()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(f"SELECT {PUBLIC_USER_FIELDS} FROM users WHERE id=%s", (user_id,))
            user = cur.fetchone()
    finally:
        conn.close()
    if not user or user["status"] != "active":
        request.session.clear()
        raise HTTPException(401, "로그인이 필요합니다.")
    return dict(user)


def require_admin(user: dict = Depends(get_current_user)) -> dict:
    if user["role"] != "admin":
        raise HTTPException(403, "관리자 권한이 필요합니다.")
    return user


class SignupRequest(BaseModel):
    name: str
    email: str
    password: str


class LoginRequest(BaseModel):
    email: str
    password: str


class ChangePasswordRequest(BaseModel):
    current_password: str
    new_password: str


class AdminUserCreateRequest(BaseModel):
    name: str
    email: str
    password: str
    role: str = "user"
    status: str = "active"


class AdminUserUpdateRequest(BaseModel):
    name: str | None = None
    email: str | None = None
    role: str | None = None
    password: str | None = None
    status: str | None = None


@app.post("/api/auth/signup")
def signup(req: SignupRequest):
    name = req.name.strip()
    email = req.email.strip().lower()
    if not name or not email or len(req.password) < 4:
        raise HTTPException(400, "이름, 이메일, 4자 이상의 비밀번호를 입력해주세요.")
    conn = get_users_conn()
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT 1 FROM users WHERE email=%s", (email,))
            if cur.fetchone():
                raise HTTPException(400, "이미 가입된 이메일입니다.")
            cur.execute(
                "INSERT INTO users (name, email, password_hash, role, status) VALUES (%s,%s,%s,'user','active')",
                (name, email, hash_password(req.password)),
            )
        conn.commit()
    finally:
        conn.close()
    return {"ok": True}


@app.post("/api/auth/login")
def login(req: LoginRequest, request: Request):
    conn = get_users_conn()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("SELECT * FROM users WHERE email=%s", (req.email.strip().lower(),))
            user = cur.fetchone()
    finally:
        conn.close()
    if not user or not verify_password(req.password, user["password_hash"]):
        raise HTTPException(401, "이메일 또는 비밀번호가 올바르지 않습니다.")
    if user["status"] != "active":
        raise HTTPException(403, "비활성화된 계정입니다. 관리자에게 문의하세요.")
    request.session["user_id"] = user["id"]
    return {
        "id": user["id"], "name": user["name"], "email": user["email"],
        "role": user["role"], "created_at": user["created_at"].isoformat(),
    }


@app.post("/api/auth/logout")
def logout(request: Request):
    request.session.clear()
    return {"ok": True}


@app.get("/api/auth/me")
def me(user: dict = Depends(get_current_user)):
    return {
        "id": user["id"], "name": user["name"], "email": user["email"],
        "role": user["role"], "created_at": user["created_at"].isoformat(),
    }


@app.put("/api/auth/me/password")
def change_my_password(req: ChangePasswordRequest, user: dict = Depends(get_current_user)):
    if len(req.new_password) < 4:
        raise HTTPException(400, "새 비밀번호는 4자 이상이어야 합니다.")
    conn = get_users_conn()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("SELECT password_hash FROM users WHERE id=%s", (user["id"],))
            row = cur.fetchone()
            if not row or not verify_password(req.current_password, row["password_hash"]):
                raise HTTPException(400, "현재 비밀번호가 올바르지 않습니다.")
            cur.execute("UPDATE users SET password_hash=%s WHERE id=%s", (hash_password(req.new_password), user["id"]))
        conn.commit()
    finally:
        conn.close()
    return {"ok": True}


@app.get("/api/admin/users")
def admin_list_users(admin: dict = Depends(require_admin)):
    conn = get_users_conn()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(f"SELECT {PUBLIC_USER_FIELDS} FROM users ORDER BY created_at")
            rows = cur.fetchall()
    finally:
        conn.close()
    return [
        {**dict(r), "created_at": r["created_at"].isoformat(), "status_changed_at": r["status_changed_at"].isoformat()}
        for r in rows
    ]


@app.post("/api/admin/users")
def admin_create_user(req: AdminUserCreateRequest, admin: dict = Depends(require_admin)):
    name = req.name.strip()
    email = req.email.strip().lower()
    if not name or not email or len(req.password) < 4:
        raise HTTPException(400, "이름, 이메일, 4자 이상의 비밀번호를 입력해주세요.")
    if req.role not in ("user", "admin") or req.status not in ("active", "inactive"):
        raise HTTPException(400, "잘못된 값입니다.")
    conn = get_users_conn()
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT 1 FROM users WHERE email=%s", (email,))
            if cur.fetchone():
                raise HTTPException(400, "이미 존재하는 이메일입니다.")
            cur.execute(
                "INSERT INTO users (name, email, password_hash, role, status) VALUES (%s,%s,%s,%s,%s)",
                (name, email, hash_password(req.password), req.role, req.status),
            )
        conn.commit()
    finally:
        conn.close()
    return {"ok": True}


@app.put("/api/admin/users/{user_id}")
def admin_update_user(user_id: int, req: AdminUserUpdateRequest, admin: dict = Depends(require_admin)):
    fields, params = [], []
    if req.name is not None:
        fields.append("name=%s")
        params.append(req.name.strip())
    if req.email is not None:
        fields.append("email=%s")
        params.append(req.email.strip().lower())
    if req.role is not None:
        if req.role not in ("user", "admin"):
            raise HTTPException(400, "잘못된 권한 값입니다.")
        fields.append("role=%s")
        params.append(req.role)
    if req.password:
        if len(req.password) < 4:
            raise HTTPException(400, "비밀번호는 4자 이상이어야 합니다.")
        fields.append("password_hash=%s")
        params.append(hash_password(req.password))
    if req.status is not None:
        if req.status not in ("active", "inactive"):
            raise HTTPException(400, "잘못된 상태 값입니다.")
        fields.append("status_changed_at=CASE WHEN status<>%s THEN now() ELSE status_changed_at END")
        params.append(req.status)
        fields.append("status=%s")
        params.append(req.status)
    if not fields:
        return {"ok": True}
    params.append(user_id)
    conn = get_users_conn()
    try:
        with conn.cursor() as cur:
            cur.execute(f"UPDATE users SET {', '.join(fields)} WHERE id=%s", params)
        conn.commit()
    finally:
        conn.close()
    return {"ok": True}


@app.delete("/api/admin/users/{user_id}")
def admin_delete_user(user_id: int, admin: dict = Depends(require_admin)):
    if user_id == admin["id"]:
        raise HTTPException(400, "본인 계정은 삭제할 수 없습니다.")
    conn = get_users_conn()
    try:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM users WHERE id=%s", (user_id,))
        conn.commit()
    finally:
        conn.close()
    return {"ok": True}


def get_openrouter_client() -> OpenAI:
    api_key = os.environ.get("OPENROUTER_API_KEY")
    if not api_key:
        raise HTTPException(500, "OPENROUTER_API_KEY가 서버에 설정되어 있지 않습니다.")
    return OpenAI(base_url="https://openrouter.ai/api/v1", api_key=api_key)

CHAT_SYSTEM_TEMPLATE = """당신은 'US TV Market — Brand Competition Dashboard'의 데이터 분석 어시스턴트입니다.
아래는 2024~2026년 Amazon.com TV 카테고리의 주차별 브랜드 판매 데이터(CSV)입니다.
week 컬럼은 YYYYWW 형식입니다 (예: 202520 = 2025년 20주차). conv(전환율, %)는 units/traffic*100 입니다.
브랜드 목록: Samsung, LG, TCL, Insignia, HiSense, Amazon, Sony, Roku, Vizio, Toshiba.

답변은 반드시 아래 JSON 스키마를 따르는 하나의 JSON 객체로만 출력하세요. JSON 앞뒤에 다른 텍스트를 붙이지 마세요.

{{
  "blocks": [
    {{"type": "text", "content": "설명 문장", "source": null}},
    {{"type": "table", "title": "표 제목(선택, 없으면 빈 문자열)", "headers": ["열1", "열2"], "rows": [["값1", "값2"]], "source": null}},
    {{"type": "chart", "chartType": "bar 또는 line", "title": "차트 제목(선택)", "xKey": "label", "series": [{{"key": "value", "name": "계열 이름"}}], "data": [{{"label": "202519", "value": 12345}}], "source": null}}
  ],
  "suggestions": ["팔로우업 질문1", "팔로우업 질문2", "팔로우업 질문3"]
}}

블록 작성 규칙:
1. 매출/트래픽/판매량/전환율 등 데이터에서 직접 확인 가능한 수치는 CSV를 근거로 정확히 계산해서 답하세요.
2. "왜 ~했는지", "원인이 뭐야" 같은 질문에는 먼저 CSV 데이터 안에서 근거(전후 주차 대비 변화, 다른 브랜드와의 비교, 계절성 등)를 찾아 분석하고,
   검색으로 얻은 최신 뉴스/기사(신제품 출시, 프로모션, 이벤트, 시즌 세일 등)가 있다면 근거로 함께 활용하세요.
   URL이나 도메인 이름(예: example.com)은 content 문자열 안에 절대 직접 쓰지 마세요 — "[텍스트](URL)" 마크다운 링크 문법도 금지입니다.
   출처는 "OO 매체에 따르면"처럼 자연스러운 문장으로만 언급하고, 실제 URL은 반드시 아래 8번 규칙대로 source 필드에만 넣으세요.
3. 여러 주차나 여러 브랜드에 걸친 수치 비교/추이는 text로 나열하지 말고 table 또는 chart 블록으로 표현해서 가독성을 높이세요.
   시간 흐름(주차별 추이)에는 chart(line)를, 브랜드 간 비교에는 chart(bar) 또는 table을 우선 사용하세요.
   설명, 원인 분석, 결론처럼 표/차트로 표현하기 어려운 내용은 text 블록으로 쓰세요.
   한 답변 안에 여러 블록을 섞어 써도 됩니다 (예: text로 요약 → table/chart로 세부 수치 → text로 결론).
4. chart 블록의 data 안 숫자 값(value 등 series의 key에 해당하는 값)은 반드시 순수 숫자로 넣으세요 (문자열 금지).
5. 답변은 간결한 한국어로 작성하고, 확실한 사실과 추정을 구분해서 표현하세요 (예: "~로 추정됩니다").
6. text/table 문자열 안에서 사용자가 꼭 봐야 할 핵심 수치나 결론(최고/최저 값, 핵심 인사이트 등)은 **핵심내용**처럼 별표 두 개로 감싸서 강조하세요.
   이 강조 표시는 채팅창에서 초록색 굵은 글씨로 표시됩니다. 문장 전체를 감싸지 말고, 정말 중요한 숫자나 결론 한둘만 강조하세요.
   그 외 #, 별표 하나(*), -, |, `, > 같은 마크다운 기호나 이모지는 절대 쓰지 마세요.
7. suggestions에는 이번 답변과 지금까지의 대화 맥락을 고려했을 때 사용자가 다음으로 물어보면 유용할 구체적인 질문 3개를 만드세요.
   데이터로 답할 수 있거나 추가 분석이 가능한, 실용적인 질문으로 작성하고, 이미 방금 답한 내용을 그대로 반복하는 질문은 피하세요.
8. 각 블록의 source 필드에는 그 블록 내용의 근거를 표시하세요. CSV 데이터에서 나온 내용이면 source는 null로 두세요(화면에 아무것도 표시되지 않습니다).
   웹 검색으로 찾은 내용이면 실제로 참고한 기사/페이지의 URL 하나를 source에 넣으세요. 여러 출처를 참고했다면 가장 핵심적인 URL 하나만 넣으세요.
   CSV 데이터에서 나온 내용과 웹 검색으로 찾은 내용을 하나의 블록에 섞지 마세요 — 출처가 다르면 반드시 별도의 text 블록으로 나누고, 각 블록에 맞는 source를 넣으세요.

데이터(CSV):
{data_csv}
"""


class ChatMessage(BaseModel):
    role: str
    content: str


class ChatRequest(BaseModel):
    message: str
    history: list[ChatMessage] = []


def strip_markdown(text: str) -> str:
    # NOTE: **bold** is intentionally left intact — the frontend renders it as
    # a green highlight span (see rule 6 in CHAT_SYSTEM_TEMPLATE). Every other
    # markdown construct is stripped since the chat panel shows raw text.
    text = re.sub(r'\[([^\]]+)\]\(([^)]+)\)', r'\1 (\2)', text)
    text = re.sub(r'__(.+?)__', r'\1', text)
    text = re.sub(r'(?<!\*)\*([^\*\n]+?)\*(?!\*)', r'\1', text)
    text = re.sub(r'`([^`]+)`', r'\1', text)
    text = re.sub(r'^\s{0,3}#{1,6}\s*', '', text, flags=re.MULTILINE)
    text = re.sub(r'^\s{0,3}>\s?', '', text, flags=re.MULTILINE)
    text = re.sub(r'^\s{0,3}[-*+]\s+', '', text, flags=re.MULTILINE)
    return text


DEFAULT_SUGGESTIONS = [
    "브랜드별 매출과 전환율을 표로 비교해줘",
    "최근 3개월간 삼성과 LG의 매출 추이를 차트로 보여줘",
    "전환율이 가장 낮은 브랜드는 어디이고, 개선하려면 어떤 점을 봐야 할까?",
]


def clean_source(raw_source) -> str | None:
    if not isinstance(raw_source, str):
        return None
    s = raw_source.strip()
    return s if re.match(r'^https?://', s) else None


URL_RE = re.compile(r'https?://[^\s)\]]+')
# Common leftover pattern from a markdown-link conversion: "kpinews.kr (https://...)"
# -- a bare domain-name mention sitting right before the URL it names.
DOMAIN_THEN_URL_RE = re.compile(r'\S{1,30}\.[a-zA-Z]{2,6}\s*\(\s*(https?://[^\s)\]]+)\s*\)')


def extract_and_strip_urls(text: str):
    """Safety net for when the model writes a URL inline instead of using the
    block's `source` field (e.g. leftover from a markdown-link conversion).
    Pulls the first URL out to use as the source and removes all URL text --
    plus a bare domain-name mention immediately preceding it, if present --
    from the visible content."""
    urls = []

    def _domain_sub(m):
        urls.append(m.group(1))
        return ''

    cleaned = DOMAIN_THEN_URL_RE.sub(_domain_sub, text)
    bare_urls = URL_RE.findall(cleaned)
    if bare_urls:
        urls.extend(bare_urls)
        cleaned = URL_RE.sub('', cleaned)
    if not urls:
        return text, None
    cleaned = re.sub(r'\(\s*\)', '', cleaned)
    cleaned = re.sub(r'[ \t]+([,.!?])', r'\1', cleaned)
    cleaned = re.sub(r'[ \t]{2,}', ' ', cleaned)
    cleaned = re.sub(r'\n{3,}', '\n\n', cleaned)
    return cleaned.strip(), urls[0]


def url_is_reachable(url: str, timeout: float = 4.0) -> bool:
    """LLM-cited URLs are sometimes reconstructed/hallucinated rather than copied
    verbatim from search results, so a link-icon can point at a 404. Verify live
    before showing it rather than trusting the model's citation blindly."""
    try:
        with httpx.Client(follow_redirects=True, timeout=timeout) as client:
            resp = client.head(url)
            if resp.status_code == 405:  # some sites reject HEAD; retry with GET
                resp = client.get(url)
            return resp.status_code < 400
    except Exception:
        return False


def drop_dead_sources(blocks: list) -> list:
    cache: dict[str, bool] = {}
    for b in blocks:
        url = b.get("source")
        if not url:
            continue
        if url not in cache:
            cache[url] = url_is_reachable(url)
        if not cache[url]:
            b["source"] = None
    return blocks


def parse_chat_response(raw: str):
    try:
        obj = json.loads(raw)
    except Exception:
        content, extracted_url = extract_and_strip_urls(strip_markdown(raw or ""))
        return [{"type": "text", "content": content, "source": extracted_url}], DEFAULT_SUGGESTIONS

    blocks = obj.get("blocks")
    if not isinstance(blocks, list):
        blocks = []

    cleaned = []
    for b in blocks:
        if not isinstance(b, dict):
            continue
        btype = b.get("type")
        source = clean_source(b.get("source"))
        if btype == "text":
            content = strip_markdown(str(b.get("content", "")))
            content, extracted_url = extract_and_strip_urls(content)
            source = source or extracted_url
            if content.strip():
                cleaned.append({"type": "text", "content": content, "source": source})
        elif btype == "table":
            headers = [strip_markdown(str(h)) for h in (b.get("headers") or [])]
            rows = [[strip_markdown(str(c)) for c in row] for row in (b.get("rows") or [])]
            if headers and rows:
                cleaned.append({
                    "type": "table",
                    "title": strip_markdown(str(b.get("title") or "")),
                    "headers": headers,
                    "rows": rows,
                    "source": source,
                })
        elif btype == "chart":
            data = b.get("data")
            series = b.get("series") or [{"key": "value", "name": "값"}]
            if isinstance(data, list) and data:
                cleaned.append({
                    "type": "chart",
                    "chartType": b.get("chartType") if b.get("chartType") in ("bar", "line") else "bar",
                    "title": strip_markdown(str(b.get("title") or "")),
                    "xKey": b.get("xKey") or "label",
                    "series": series,
                    "data": data,
                    "source": source,
                })
    cleaned = cleaned or [{"type": "text", "content": "답변을 생성하지 못했습니다.", "source": None}]
    cleaned = drop_dead_sources(cleaned)

    suggestions = []
    for s in (obj.get("suggestions") or []):
        if isinstance(s, str) and s.strip():
            suggestions.append(strip_markdown(s.strip()))
    suggestions = suggestions[:3] or DEFAULT_SUGGESTIONS

    return cleaned, suggestions


def build_data_csv() -> str:
    conn = get_conn()
    rows = conn.execute(
        "SELECT week_id AS week, brand, rev, units, traffic FROM weekly_sales ORDER BY week_id, brand"
    ).fetchall()
    conn.close()
    lines = ["week,brand,rev,units,traffic"]
    for r in rows:
        lines.append(f"{r['week']},{r['brand']},{r['rev']},{r['units']},{r['traffic']}")
    return "\n".join(lines)


def get_conn():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


@app.get("/api/weekly")
def weekly(user: dict = Depends(get_current_user)):
    conn = get_conn()
    rows = conn.execute(
        "SELECT week_id AS week, brand, rev, units, traffic, organic FROM weekly_sales ORDER BY week_id, brand"
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


@app.get("/api/totals")
def totals(user: dict = Depends(get_current_user)):
    conn = get_conn()
    rows = conn.execute(
        """
        SELECT brand,
               SUM(rev) AS rev,
               SUM(units) AS units,
               SUM(traffic) AS traffic,
               SUM(organic) AS organic
        FROM weekly_sales
        GROUP BY brand
        """
    ).fetchall()
    conn.close()
    result = []
    for r in rows:
        conv = (r["units"] / r["traffic"] * 100) if r["traffic"] else 0
        result.append({
            "brand": r["brand"],
            "rev": r["rev"],
            "units": r["units"],
            "traffic": r["traffic"],
            "organic": r["organic"],
            "conv": round(conv, 2),
        })
    return result


def _display_width(s: str) -> float:
    """Approximate rendered width of a string in Excel column-width units.
    Hangul/CJK characters render roughly twice as wide as Latin/digit
    characters, so a naive len() undercounts Korean-heavy cells."""
    width = 0.0
    for ch in s:
        code = ord(ch)
        if 0x1100 <= code <= 0x115F or 0x2E80 <= code <= 0xA4CF or 0xAC00 <= code <= 0xD7A3 or 0xF900 <= code <= 0xFAFF or 0xFF00 <= code <= 0xFF60 or 0xFFE0 <= code <= 0xFFE6:
            width += 2
        else:
            width += 1
    return width


def _parse_week(week: str, field_name: str) -> int:
    if not re.fullmatch(r"\d{1,2}", week):
        raise HTTPException(400, f"잘못된 {field_name} 값입니다.")
    return int(week)


def period_where(period: str, week_from: str = "all", week_to: str = "all"):
    """Returns (sql_clause, params) filtering sku_weekly.week_id to a year and
    optionally a week range within it, or ('', []) for 'all'."""
    if period == "all":
        return "", []
    if not re.fullmatch(r"\d{4}", period):
        raise HTTPException(400, "잘못된 period 값입니다.")
    if week_from == "all":
        return "WHERE week_id / 100 = ?", [int(period)]
    wf = _parse_week(week_from, "week_from")
    wt = _parse_week(week_to, "week_to") if week_to != "all" else wf
    return "WHERE week_id / 100 = ? AND week_id % 100 BETWEEN ? AND ?", [int(period), wf, wt]


OS_LABELS = {"os_-": "기타/미상"}

TOP_MODELS_SORT_COLUMNS = {
    "units_sold": "units_sold",
    "retail_sales": "retail_sales",
    "conversion_rate": "conversion_rate",
    "total_traffic": "total_traffic",
    "organic_traffic": "organic_traffic",
}


@app.get("/api/os_share")
def os_share(period: str = "all", week_from: str = "all", week_to: str = "all", user: dict = Depends(get_current_user)):
    where_clause, params = period_where(period, week_from, week_to)
    conn = get_conn()
    rows = conn.execute(
        f"""
        SELECT os, SUM(units_sold) AS units
        FROM sku_weekly
        {where_clause}
        GROUP BY os
        ORDER BY units DESC
        """,
        params,
    ).fetchall()
    conn.close()
    return [{"os": OS_LABELS.get(r["os"], r["os"] or "기타/미상"), "units": r["units"]} for r in rows]


@app.get("/api/top_models")
def top_models(period: str = "all", week_from: str = "all", week_to: str = "all", sort: str = "units_sold", limit: int = 15, user: dict = Depends(get_current_user)):
    sort_col = TOP_MODELS_SORT_COLUMNS.get(sort)
    if sort_col is None:
        raise HTTPException(400, "잘못된 sort 값입니다.")
    limit = max(1, min(limit, 100))
    where_clause, params = period_where(period, week_from, week_to)

    conn = get_conn()
    rows = conn.execute(
        f"""
        SELECT
            MAX(brand) AS brand,
            retailer_sku,
            MAX(model_number) AS model_number,
            MAX(os) AS os,
            SUM(units_sold) AS units_sold,
            SUM(retail_sales) AS retail_sales,
            SUM(total_traffic) AS total_traffic,
            SUM(organic_traffic) AS organic_traffic,
            CASE WHEN SUM(total_traffic) > 0
                 THEN SUM(units_sold) * 100.0 / SUM(total_traffic)
                 ELSE 0 END AS conversion_rate
        FROM sku_weekly
        {where_clause}
        GROUP BY retailer_sku
        ORDER BY {sort_col} DESC
        LIMIT ?
        """,
        params + [limit],
    ).fetchall()
    conn.close()
    return [
        {
            "brand": r["brand"],
            "retailer_sku": r["retailer_sku"],
            "model_number": r["model_number"] or "-",
            "units_sold": r["units_sold"],
            "retail_sales": round(r["retail_sales"], 2),
            "conversion_rate": round(r["conversion_rate"], 2),
            "total_traffic": r["total_traffic"],
            "organic_traffic": r["organic_traffic"],
            "os": OS_LABELS.get(r["os"], r["os"] or "기타/미상"),
        }
        for r in rows
    ]


@app.get("/api/export")
def export_raw_data(period: str = "all", week_from: str = "all", week_to: str = "all", user: dict = Depends(get_current_user)):
    where_clause, params = period_where(period, week_from, week_to)
    conn = get_conn()
    rows = conn.execute(
        f"""
        SELECT week_id, brand, rev, units, traffic, organic
        FROM weekly_sales
        {where_clause}
        ORDER BY week_id, brand
        """,
        params,
    ).fetchall()
    conn.close()

    year_label = "전체" if period == "all" else f"{period}년"
    from_label = "전체" if week_from == "all" else f"{week_from}주차"
    to_label = "전체" if week_to == "all" else f"{week_to}주차"
    period_label = "전체" if period == "all" else f"{year_label} {from_label}~{to_label}"
    period_filename_label = f"{year_label} {from_label}-{to_label}"

    wb = Workbook()
    ws = wb.active
    ws.title = "Raw Data"
    bold = Font(bold=True)
    label_fill = PatternFill(start_color="595959", end_color="595959", fill_type="solid")
    label_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="000000")
    medium = Side(style="medium", color="000000")

    italic = Font(italic=True)
    summary = [("추출기간", period_label), ("추출일시", datetime.now().strftime("%Y.%m.%d %H:%M")), ("추출건수", len(rows))]
    for r, (label, value) in enumerate(summary, start=1):
        label_cell = ws.cell(row=r, column=1, value=label)
        label_cell.font = label_font
        label_cell.fill = label_fill
        label_cell.border = Border(left=None, right=thin, top=thin, bottom=thin)
        label_cell.alignment = Alignment(horizontal="center", vertical="center")
        value_cell = ws.cell(row=r, column=2, value=value)
        value_cell.font = italic
        value_cell.border = Border(left=thin, right=None, top=thin, bottom=thin)

    header_row = 5
    header_fill = PatternFill(start_color="C7C7C7", end_color="C7C7C7", fill_type="solid")
    for col_idx, h in enumerate(["주차", "브랜드", "매출", "판매수량", "트래픽", "Organic 트래픽"], start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=h)
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for i, r in enumerate(rows, start=header_row + 1):
        ws.cell(row=i, column=1, value=r["week_id"]).alignment = Alignment(horizontal="center")
        ws.cell(row=i, column=2, value=r["brand"]).alignment = Alignment(horizontal="left")
        for col_idx, key in enumerate(["rev", "units", "traffic", "organic"], start=3):
            cell = ws.cell(row=i, column=col_idx, value=r[key])
            cell.alignment = Alignment(horizontal="right")
            cell.number_format = "#,##0"

    last_row = header_row + len(rows)
    n_cols = 6
    for r in range(header_row, last_row + 1):
        for c in range(1, n_cols + 1):
            ws.cell(row=r, column=c).border = Border(
                top=medium if r == header_row else thin,
                bottom=medium if r in (header_row, last_row) else thin,
                left=None if c == 1 else thin,
                right=None if c == n_cols else thin,
            )

    for c in range(1, n_cols + 1):
        values = (ws.cell(row=r, column=c).value for r in range(1, last_row + 1))
        width = max((_display_width(str(v)) for v in values if v is not None), default=8) + 3
        if c == 1:
            width *= 4 / 3
        ws.column_dimensions[get_column_letter(c)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"Raw Data_{period_filename_label}_{datetime.now().strftime('%Y%m%d %H%M')}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
    )


def call_chat_model(client: OpenAI, messages: list):
    attempts = [
        dict(model=CHAT_MODEL, messages=messages, response_format={"type": "json_object"}),
        dict(model=CHAT_MODEL, messages=messages),
    ]
    last_err = None
    for kwargs in attempts:
        try:
            return client.chat.completions.create(**kwargs)
        except Exception as e:
            last_err = e
    raise last_err


@app.post("/api/chat")
def chat(req: ChatRequest, user: dict = Depends(get_current_user)):
    client = get_openrouter_client()

    system = CHAT_SYSTEM_TEMPLATE.format(data_csv=build_data_csv())
    messages = [{"role": "system", "content": system}]

    if needs_external_search(req.message):
        naver_context = build_naver_context(req.message)
        if naver_context:
            messages.append({
                "role": "system",
                "content": "다음은 네이버 뉴스/블로그 실시간 검색 결과입니다. 질문과 관련 있는 내용이 있으면 "
                            "답변의 근거로 활용하고, 해당 블록의 source 필드에 정확한 URL을 넣으세요. 관련 없으면 무시하세요:\n"
                            + naver_context,
            })

    messages += [{"role": m.role, "content": m.content} for m in req.history]
    messages.append({"role": "user", "content": req.message})

    try:
        resp = call_chat_model(client, messages)
    except Exception as e:
        print(f"chat error: {e!r}")
        raise HTTPException(500, "AI 응답 생성 중 오류가 발생했습니다. 잠시 후 다시 시도해주세요.")

    raw = resp.choices[0].message.content
    blocks, suggestions = parse_chat_response(raw)
    return {"blocks": blocks, "suggestions": suggestions}


@app.get("/")
def index():
    return FileResponse(ROOT / "brand-dashboard.html")


app.mount("/static", StaticFiles(directory=str(ROOT)), name="static")
